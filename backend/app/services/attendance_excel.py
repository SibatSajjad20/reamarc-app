"""
Multi-Tab Excel (.xlsx) Export Generator for Attendance, Shifts & Monthly Punctuality.
Uses OpenPyXL to construct styled corporate spreadsheet reports:
- Tab 1: "Punctuality Summary" (Company-wide monthly aggregation, scores & bonus recommendations).
- Tabs 2 to N: Individual Employee Monthly Timesheets (Daily breakdown with color-coded status styling).
"""
import io
import re
import calendar
from datetime import datetime, date, timezone
from zoneinfo import ZoneInfo
from typing import Optional, List, Dict, Any, Set
from collections import defaultdict
import logging

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database import get_database
from app.models.attendance import (
    AttendanceStatus,
    LeaveType,
    LeaveStatus,
    BonusRecommendation,
    CalendarEventType,
)
from app.schemas.shift import ShiftResponse
from app.services.attendance_calculator import (
    calculate_daily_attendance,
    calculate_monthly_aggregation,
    format_minutes_to_hhmm,
)
from app.services import attendance_service

logger = logging.getLogger("app.attendance_excel")

PK_TZ = ZoneInfo("Asia/Karachi")

# ──────────────────────────────────────────────────────────
# STYLING DEFINITIONS & PALETTE
# ──────────────────────────────────────────────────────────
NAVY_HEADER_HEX = "1E293B"       # Slate 800
WHITE_HEX = "FFFFFF"
ALT_ROW_HEX = "F8FAFC"          # Slate 50
YELLOW_WEEKEND_HEX = "FEF3C7"   # Amber 100 (Weekends & Holidays)
RED_LATE_HEX = "FEE2E2"         # Red 100 (Late badge fill)
BLUE_WFH_HEX = "DBEAFE"         # Blue 100 (WFH & Leaves)
FOOTER_GRAY_HEX = "E2E8F0"      # Slate 200 (Footer summary)
BORDER_GRAY_HEX = "CBD5E1"      # Slate 300
BORDER_DARK_HEX = "0F172A"      # Slate 900

FONT_NAME = "Calibri"

# Fonts
FONT_TITLE = Font(name=FONT_NAME, size=13, bold=True, color="1E293B")
FONT_SUBTITLE = Font(name=FONT_NAME, size=10, italic=True, color="64748B")
FONT_HEADER = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
FONT_DATA_REGULAR = Font(name=FONT_NAME, size=10, color="0F172A")
FONT_DATA_BOLD = Font(name=FONT_NAME, size=10, bold=True, color="0F172A")
FONT_DATA_MUTED = Font(name=FONT_NAME, size=10, color="64748B")
FONT_OVERTIME = Font(name=FONT_NAME, size=10, bold=True, italic=True, color="16A34A")  # Bold Green Italic
FONT_UNDERTIME = Font(name=FONT_NAME, size=10, bold=True, color="DC2626")             # Bold Red
FONT_LATE = Font(name=FONT_NAME, size=10, bold=True, color="DC2626")                  # Bold Red
FONT_BONUS_ELIGIBLE = Font(name=FONT_NAME, size=10, bold=True, color="16A34A")        # Green
FONT_BONUS_REVIEW = Font(name=FONT_NAME, size=10, bold=True, color="D97706")          # Amber
FONT_BONUS_NOT_ELIGIBLE = Font(name=FONT_NAME, size=10, bold=True, color="DC2626")    # Red

# Fills
FILL_NAVY = PatternFill(start_color=NAVY_HEADER_HEX, end_color=NAVY_HEADER_HEX, fill_type="solid")
FILL_WHITE = PatternFill(start_color=WHITE_HEX, end_color=WHITE_HEX, fill_type="solid")
FILL_ALT = PatternFill(start_color=ALT_ROW_HEX, end_color=ALT_ROW_HEX, fill_type="solid")
FILL_WEEKEND = PatternFill(start_color=YELLOW_WEEKEND_HEX, end_color=YELLOW_WEEKEND_HEX, fill_type="solid")
FILL_LATE_BADGE = PatternFill(start_color=RED_LATE_HEX, end_color=RED_LATE_HEX, fill_type="solid")
FILL_BLUE_WFH = PatternFill(start_color=BLUE_WFH_HEX, end_color=BLUE_WFH_HEX, fill_type="solid")
FILL_FOOTER = PatternFill(start_color=FOOTER_GRAY_HEX, end_color=FOOTER_GRAY_HEX, fill_type="solid")

# Borders
THIN_SIDE = Side(style="thin", color=BORDER_GRAY_HEX)
MEDIUM_SIDE = Side(style="medium", color=BORDER_DARK_HEX)
DOUBLE_SIDE = Side(style="double", color=BORDER_DARK_HEX)

BORDER_CELL = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
BORDER_HEADER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=MEDIUM_SIDE)
BORDER_FOOTER = Border(left=THIN_SIDE, right=THIN_SIDE, top=MEDIUM_SIDE, bottom=DOUBLE_SIDE)

# Alignments
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def sanitize_sheet_title(name: str, used_titles: Set[str]) -> str:
    """
    Sanitizes employee name for Excel worksheet titles:
    - Removes invalid characters: []:*?/\\
    - Truncates to max 31 characters.
    - Handles collisions by appending unique suffixes (_2, _3).
    """
    clean_name = re.sub(r"[\[\]:*?/\\]", "", name).strip()
    if not clean_name:
        clean_name = "Employee"

    base_title = clean_name[:28] if len(clean_name) > 28 else clean_name
    candidate = base_title[:31]

    counter = 1
    while candidate.lower() in used_titles:
        counter += 1
        suffix = f"_{counter}"
        candidate = f"{base_title[:31 - len(suffix)]}{suffix}"

    used_titles.add(candidate.lower())
    return candidate


def auto_fit_columns(ws, min_width: int = 12, padding: int = 3):
    """Calculates and sets optimal column widths for an openpyxl worksheet."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            # Skip merged cells with huge text to prevent giant column expansion
            if cell.coordinate in ws.merged_cells:
                continue
            if cell.value is not None:
                val_str = str(cell.value)
                lines = val_str.split("\n")
                line_len = max(len(l) for l in lines)
                if line_len > max_len:
                    max_len = line_len
        ws.column_dimensions[col_letter].width = max(max_len + padding, min_width)


async def generate_multi_tab_attendance_workbook(
    year: int,
    month: int,
    department: Optional[str] = None,
) -> io.BytesIO:
    """
    Generates a production-grade multi-tab .xlsx workbook containing:
    1. Tab 1: 'Punctuality Summary' (Company-wide monthly aggregation).
    2. Tabs 2..N: Individual Employee Monthly Timesheets (Detailed daily register).
    Returns an in-memory io.BytesIO stream.
    """
    db = get_database()
    month_prefix = f"{year:04d}-{month:02d}"
    month_name = calendar.month_name[month]
    num_days = calendar.monthrange(year, month)[1]

    # 1. Fetch Users (Exclude client and admin roles from employee attendance records)
    user_query: Dict[str, Any] = {"is_active": True, "role": {"$nin": ["client", "CLIENT", "admin", "ADMIN"]}}
    if department and department.lower() != "all":
        user_query["department"] = {"$regex": f"^{department}$", "$options": "i"}

    users = []
    if db is not None:
        users = await db.users.find(user_query, {"_id": 0, "hashed_password": 0}).sort("full_name", 1).to_list(1000)

    # 2. Fetch Company Calendar Events for target month
    calendar_events = []
    holidays_set: Set[str] = set()
    working_saturdays_set: Set[str] = set()
    calendar_titles: Dict[str, str] = {}

    if db is not None:
        calendar_events = await db.company_calendar.find(
            {"date": {"$regex": f"^{month_prefix}-"}},
            {"_id": 0}
        ).to_list(100)
        for ev in calendar_events:
            ev_date = ev.get("date")
            ev_type = ev.get("event_type")
            calendar_titles[ev_date] = ev.get("title") or "Event"
            if ev_type in (CalendarEventType.HOLIDAY.value, "holiday"):
                holidays_set.add(ev_date)
            if ev.get("is_workday_override") or ev_type in (CalendarEventType.WORKING_SATURDAY.value, "working_saturday"):
                working_saturdays_set.add(ev_date)

    # 3. Calculate working days in month
    total_working_days = await attendance_service.calculate_month_working_days(year, month)

    # 4. Fetch Attendance Records and Leaves for all users for this month
    records_by_user = defaultdict(list)
    records_by_user_date = defaultdict(dict)
    if db is not None:
        all_records = await db.attendance_records.find(
            {"date": {"$regex": f"^{month_prefix}-"}},
            {"_id": 0}
        ).sort("date", 1).to_list(10000)
        for r in all_records:
            u_id = r.get("user_id")
            records_by_user[u_id].append(r)
            records_by_user_date[u_id][r.get("date")] = r

    start_of_month = f"{month_prefix}-01"
    end_of_month = f"{month_prefix}-{num_days:02d}"

    leaves_by_user = defaultdict(list)
    leaves_by_user_date = defaultdict(dict)
    if db is not None:
        all_leaves = await db.leave_requests.find(
            {
                "status": LeaveStatus.APPROVED.value,
                "start_date": {"$lte": end_of_month},
                "end_date": {"$gte": start_of_month},
            },
            {"_id": 0}
        ).to_list(2000)
        for l in all_leaves:
            u_id = l.get("user_id")
            leaves_by_user[u_id].append(l)
            start_d = l.get("start_date")
            end_d = l.get("end_date")
            # Populate date range map for quick lookup
            try:
                s_dt = datetime.strptime(start_d, "%Y-%m-%d").date()
                e_dt = datetime.strptime(end_d, "%Y-%m-%d").date()
                cur_dt = s_dt
                while cur_dt <= e_dt:
                    if cur_dt.year == year and cur_dt.month == month:
                        leaves_by_user_date[u_id][cur_dt.strftime("%Y-%m-%d")] = l
                    cur_dt += timedelta(days=1)
            except Exception:
                pass

    # 5. Pre-fetch shifts and user assignments in batch for instant in-memory lookup
    all_shifts_raw = []
    assignments_by_user = {}
    if db is not None:
        all_shifts_raw = await db.shifts.find({"is_active": True}, {"_id": 0}).to_list(100)
        raw_assignments = await db.user_shift_assignments.find({}, {"_id": 0}).to_list(2000)
        for a in raw_assignments:
            if a.get("user_id"):
                assignments_by_user[a["user_id"]] = a

    shifts_by_id = {s["id"]: ShiftResponse(**s) for s in all_shifts_raw if s.get("id")}
    std_shift_raw, hr_shift_raw = attendance_service.resolve_fallback_shifts(all_shifts_raw)
    default_std_shift = ShiftResponse(**std_shift_raw) if std_shift_raw else ShiftResponse(id="shift_std", name="Standard Shift", start_time="09:30", end_time="18:30", expected_hours=8.0)
    default_hr_shift = ShiftResponse(**hr_shift_raw) if hr_shift_raw else default_std_shift

    def resolve_user_shift_fast(u_id: str, u_dept: Optional[str] = None, d_str: Optional[str] = None) -> ShiftResponse:
        t_date = d_str or f"{year}-{month:02d}-01"
        assignment = assignments_by_user.get(u_id)
        if assignment:
            resolved = attendance_service.resolve_shift_assignment_for_date(assignment, t_date)
            shift_id = resolved.get("shift_id")
            if shift_id and shift_id in shifts_by_id:
                return shifts_by_id[shift_id]
        if u_dept and str(u_dept).upper() == "HR":
            return default_hr_shift
        return default_std_shift

    # Create Workbook
    wb = openpyxl.Workbook()
    used_sheet_titles: Set[str] = set()

    # ──────────────────────────────────────────────────────────
    # TAB 1: PUNCTUALITY SUMMARY
    # ──────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Punctuality Summary"
    used_sheet_titles.add("punctuality summary")
    ws_summary.views.sheetView[0].showGridLines = True

    # Title Row 1
    ws_summary.merge_cells("A1:Q1")
    ws_summary["A1"] = f"REAMARC AI — COMPANY MONTHLY PUNCTUALITY & ATTENDANCE SUMMARY ({month_name.upper()} {year})"
    ws_summary["A1"].font = FONT_TITLE
    ws_summary["A1"].alignment = ALIGN_LEFT
    ws_summary.row_dimensions[1].height = 28

    # Subtitle Row 2
    ws_summary.merge_cells("A2:Q2")
    ws_summary["A2"] = (
        f"Generated: {datetime.now(PK_TZ).strftime('%Y-%m-%d %H:%M PKT')} | "
        f"Department: {department or 'All Departments'} | "
        f"Total Active Staff: {len(users)} | "
        f"Scheduled Workdays: {total_working_days}"
    )
    ws_summary["A2"].font = FONT_SUBTITLE
    ws_summary["A2"].alignment = ALIGN_LEFT
    ws_summary.row_dimensions[2].height = 18

    # Blank Row 3
    ws_summary.row_dimensions[3].height = 8

    # Table Headers Row 4
    summary_headers = [
        "Emp Code",
        "Employee Name",
        "Department",
        "Shift",
        "Scheduled Days",
        "Present Days",
        "Leaves",
        "Late Strikes",
        "Short Leaves",
        "Missed Punches",
        "Expected Hours",
        "Actual Hours",
        "Overtime",
        "Undertime",
        "Net Balance",
        "Punctuality Score",
        "Bonus Recommendation",
    ]

    ws_summary.row_dimensions[4].height = 28
    for col_idx, h_text in enumerate(summary_headers, start=1):
        cell = ws_summary.cell(row=4, column=col_idx, value=h_text)
        cell.font = FONT_HEADER
        cell.fill = FILL_NAVY
        cell.border = BORDER_HEADER
        cell.alignment = ALIGN_HEADER

    # Data Rows
    current_row = 5
    tot_scheduled = 0
    tot_present = 0
    tot_leaves = 0
    tot_late = 0
    tot_short = 0
    tot_missed = 0
    tot_work_mins = 0
    tot_ot_mins = 0
    tot_ut_mins = 0
    tot_net_mins = 0
    punctuality_scores_sum = 0.0

    employee_aggregates: Dict[str, Any] = {}

    for u in users:
        u_id = u.get("id")
        u_name = u.get("full_name") or u.get("name", "User")
        u_dept = u.get("department") or "General"

        shift = resolve_user_shift_fast(u_id, u_dept)
        u_records = records_by_user.get(u_id, [])
        shift_by_date: Dict[str, Any] = {}

        daily_dicts = []
        missed_punches_count = 0
        for r in u_records:
            rec_date = r.get("date")
            rec_shift = shift
            if rec_date:
                if rec_date not in shift_by_date:
                    shift_by_date[rec_date] = resolve_user_shift_fast(u_id, u_dept, rec_date)
                rec_shift = shift_by_date[rec_date]
            if not r.get("work_duration_formatted") or r.get("work_duration_formatted") == "00:00":
                attendance_service.apply_daily_calc_fields(r, rec_shift)
            st = r.get("status", AttendanceStatus.PRESENT)
            is_missed = r.get("is_missed_punch", False) or (st in (AttendanceStatus.MISSED_PUNCH.value, "missed_punch"))
            if is_missed:
                missed_punches_count += 1

            daily_dicts.append({
                "status": st,
                "late_strike": r.get("late_strike", 0),
                "work_minutes": int(r.get("working_hours_minutes") or round(float(r.get("work_hours", 0.0)) * 60)),
                "overtime_minutes": int(r.get("overtime_minutes") or round(float(r.get("overtime_hours", 0.0)) * 60)),
                "undertime_minutes": int(r.get("undertime_minutes") or round(float(r.get("undertime_hours", 0.0)) * 60)),
                "is_short_leave": (st in (AttendanceStatus.SHORT_LEAVE.value, "short_leave") or r.get("is_short_leave", False)),
            })

        agg = calculate_monthly_aggregation(daily_dicts, total_working_days=total_working_days)
        employee_aggregates[u_id] = {
            "agg": agg,
            "shift": shift,
            "shift_by_date": shift_by_date,
            "missed_punches": missed_punches_count,
        }

        # Expected hours
        expected_monthly_hours = round(total_working_days * float(shift.expected_hours), 2)

        # Totals accumulation
        tot_scheduled += agg.total_working_days
        tot_present += agg.days_present
        tot_leaves += agg.leave_count
        tot_late += agg.late_count
        tot_short += agg.short_leaves_count
        tot_missed += missed_punches_count
        tot_work_mins += agg.total_work_minutes
        tot_ot_mins += agg.overtime_minutes
        tot_ut_mins += agg.undertime_minutes
        tot_net_mins += agg.net_variance_minutes
        punctuality_scores_sum += agg.punctuality_percentage

        row_fill = FILL_WHITE if (current_row % 2 == 1) else FILL_ALT

        bonus_rec_str = agg.bonus_recommendation.value if hasattr(agg.bonus_recommendation, "value") else str(agg.bonus_recommendation)
        if bonus_rec_str == BonusRecommendation.ELIGIBLE.value or bonus_rec_str == "Eligible":
            bonus_font = FONT_BONUS_ELIGIBLE
        elif bonus_rec_str == BonusRecommendation.REVIEW.value or bonus_rec_str == "Under Review":
            bonus_font = FONT_BONUS_REVIEW
        else:
            bonus_font = FONT_BONUS_NOT_ELIGIBLE

        row_values = [
            (u_id, ALIGN_CENTER, FONT_DATA_REGULAR),
            (u_name, ALIGN_LEFT, FONT_DATA_BOLD),
            (u_dept, ALIGN_LEFT, FONT_DATA_REGULAR),
            (shift.name, ALIGN_LEFT, FONT_DATA_REGULAR),
            (agg.total_working_days, ALIGN_CENTER, FONT_DATA_REGULAR),
            (agg.days_present, ALIGN_CENTER, FONT_DATA_REGULAR),
            (agg.leave_count, ALIGN_CENTER, FONT_DATA_REGULAR),
            (agg.late_count, ALIGN_CENTER, FONT_LATE if agg.late_count > 0 else FONT_DATA_REGULAR),
            (agg.short_leaves_count, ALIGN_CENTER, FONT_DATA_REGULAR),
            (missed_punches_count, ALIGN_CENTER, FONT_LATE if missed_punches_count > 0 else FONT_DATA_REGULAR),
            (f"{expected_monthly_hours:.1f}h", ALIGN_RIGHT, FONT_DATA_REGULAR),
            (agg.total_work_hours_formatted, ALIGN_RIGHT, FONT_DATA_BOLD),
            (agg.overtime_formatted, ALIGN_RIGHT, FONT_OVERTIME if agg.overtime_minutes > 0 else FONT_DATA_MUTED),
            (agg.undertime_formatted, ALIGN_RIGHT, FONT_UNDERTIME if agg.undertime_minutes > 0 else FONT_DATA_MUTED),
            (agg.net_variance_formatted, ALIGN_RIGHT, FONT_OVERTIME if agg.net_variance_minutes >= 0 else FONT_UNDERTIME),
            (f"{agg.punctuality_percentage:.1f}%", ALIGN_CENTER, FONT_DATA_BOLD),
            (bonus_rec_str, ALIGN_CENTER, bonus_font),
        ]

        ws_summary.row_dimensions[current_row].height = 20
        for col_idx, (val, align, font) in enumerate(row_values, start=1):
            c = ws_summary.cell(row=current_row, column=col_idx, value=val)
            c.font = font
            c.fill = row_fill
            c.border = BORDER_CELL
            c.alignment = align

        current_row += 1

    # Company Footer Summary Row
    avg_score = round(punctuality_scores_sum / len(users), 1) if users else 100.0
    tot_work_formatted = format_minutes_to_hhmm(tot_work_mins, show_sign=False)
    tot_ot_formatted = format_minutes_to_hhmm(tot_ot_mins, show_sign=True)
    tot_ut_formatted = format_minutes_to_hhmm(-tot_ut_mins, show_sign=True) if tot_ut_mins > 0 else "-00:00"
    tot_net_formatted = format_minutes_to_hhmm(tot_net_mins, show_sign=True)

    ws_summary.row_dimensions[current_row].height = 24

    ws_summary.cell(row=current_row, column=1, value="TOTALS").font = FONT_DATA_BOLD
    ws_summary.cell(row=current_row, column=2, value=f"{len(users)} Employees").font = FONT_DATA_BOLD
    ws_summary.cell(row=current_row, column=3, value="-").font = FONT_DATA_BOLD
    ws_summary.cell(row=current_row, column=4, value="-").font = FONT_DATA_BOLD
    ws_summary.cell(row=current_row, column=5, value=tot_scheduled).font = FONT_DATA_BOLD
    ws_summary.cell(row=current_row, column=6, value=tot_present).font = FONT_DATA_BOLD
    ws_summary.cell(row=current_row, column=7, value=tot_leaves).font = FONT_DATA_BOLD
    ws_summary.cell(row=current_row, column=8, value=tot_late).font = FONT_DATA_BOLD
    ws_summary.cell(row=current_row, column=9, value=tot_short).font = FONT_DATA_BOLD
    ws_summary.cell(row=current_row, column=10, value=tot_missed).font = FONT_DATA_BOLD
    ws_summary.cell(row=current_row, column=11, value="-").font = FONT_DATA_BOLD
    ws_summary.cell(row=current_row, column=12, value=tot_work_formatted).font = FONT_DATA_BOLD
    ws_summary.cell(row=current_row, column=13, value=tot_ot_formatted).font = FONT_OVERTIME
    ws_summary.cell(row=current_row, column=14, value=tot_ut_formatted).font = FONT_UNDERTIME
    ws_summary.cell(row=current_row, column=15, value=tot_net_formatted).font = FONT_DATA_BOLD
    ws_summary.cell(row=current_row, column=16, value=f"{avg_score:.1f}% Avg").font = FONT_DATA_BOLD
    ws_summary.cell(row=current_row, column=17, value="-").font = FONT_DATA_BOLD

    for col_idx in range(1, 18):
        c = ws_summary.cell(row=current_row, column=col_idx)
        c.fill = FILL_FOOTER
        c.border = BORDER_FOOTER
        c.alignment = ALIGN_CENTER if col_idx not in [2, 11, 12, 13, 14, 15] else (ALIGN_LEFT if col_idx == 2 else ALIGN_RIGHT)

    auto_fit_columns(ws_summary, min_width=11, padding=3)

    # ──────────────────────────────────────────────────────────
    # TABS 2 TO N: INDIVIDUAL EMPLOYEE MONTHLY TIMESHEETS
    # ──────────────────────────────────────────────────────────
    timesheet_headers = [
        "Date",
        "Day of Week",
        "Shift",
        "Punch In",
        "Punch Out",
        "Break",
        "Working Hours",
        "Overtime",
        "Undertime",
        "Status",
        "Late Flag",
        "Remarks / Notes",
    ]

    for u in users:
        u_id = u.get("id")
        u_name = u.get("full_name") or u.get("name", "User")
        u_dept = u.get("department") or "General"

        emp_data = employee_aggregates.get(u_id, {})
        agg = emp_data.get("agg")
        shift = emp_data.get("shift") or resolve_user_shift_fast(u_id, u_dept)
        emp_shift_by_date = emp_data.get("shift_by_date", {})

        sheet_title = sanitize_sheet_title(u_name, used_sheet_titles)
        ws_emp = wb.create_sheet(title=sheet_title)
        ws_emp.views.sheetView[0].showGridLines = True

        # Header Rows 1-3
        ws_emp.merge_cells("A1:L1")
        ws_emp["A1"] = f"MONTHLY ATTENDANCE & TIMESHEET — {month_name.upper()} {year}"
        ws_emp["A1"].font = FONT_TITLE
        ws_emp["A1"].alignment = ALIGN_LEFT
        ws_emp.row_dimensions[1].height = 26

        ws_emp.merge_cells("A2:F2")
        ws_emp["A2"] = f"Employee: {u_name} (ID: {u_id}) | Department: {u_dept}"
        ws_emp["A2"].font = FONT_DATA_BOLD
        ws_emp["A2"].alignment = ALIGN_LEFT

        ws_emp.merge_cells("G2:L2")
        ws_emp["G2"] = f"Month: {month_name} {year} | Scheduled Workdays: {total_working_days}"
        ws_emp["G2"].font = FONT_DATA_REGULAR
        ws_emp["G2"].alignment = ALIGN_RIGHT
        ws_emp.row_dimensions[2].height = 18

        ws_emp.merge_cells("A3:F3")
        ws_emp["A3"] = f"Assigned Shift: {shift.name} ({shift.start_time} - {shift.end_time})"
        ws_emp["A3"].font = FONT_DATA_REGULAR
        ws_emp["A3"].alignment = ALIGN_LEFT

        ws_emp.merge_cells("G3:L3")
        ws_emp["G3"] = (
            f"Expected Work: {shift.expected_hours}h/day | "
            f"Break: {shift.break_duration_minutes}m | "
            f"Grace Buffer: {shift.grace_period_minutes}m"
        )
        ws_emp["G3"].font = FONT_DATA_REGULAR
        ws_emp["G3"].alignment = ALIGN_RIGHT
        ws_emp.row_dimensions[3].height = 18

        # Blank Row 4
        ws_emp.row_dimensions[4].height = 8

        # Timesheet Column Headers Row 5
        ws_emp.row_dimensions[5].height = 26
        for col_idx, h_text in enumerate(timesheet_headers, start=1):
            cell = ws_emp.cell(row=5, column=col_idx, value=h_text)
            cell.font = FONT_HEADER
            cell.fill = FILL_NAVY
            cell.border = BORDER_HEADER
            cell.alignment = ALIGN_HEADER

        # Daily Records Rows 6 to (6 + num_days - 1)
        u_date_records = records_by_user_date.get(u_id, {})
        u_date_leaves = leaves_by_user_date.get(u_id, {})

        emp_row = 6
        today_pkt = datetime.now(PK_TZ).date()

        for day_num in range(1, num_days + 1):
            cur_date = date(year, month, day_num)
            cur_date_str = cur_date.strftime("%Y-%m-%d")
            day_name = cur_date.strftime("%A")

            rec = u_date_records.get(cur_date_str)
            leave = u_date_leaves.get(cur_date_str)
            rec_shift = emp_shift_by_date.get(cur_date_str) or resolve_user_shift_fast(u_id, u_dept, cur_date_str)

            is_holiday = cur_date_str in holidays_set
            is_working_sat = cur_date_str in working_saturdays_set
            is_sun = (cur_date.weekday() == 6)
            is_first_sat = (cur_date.weekday() == 5 and cur_date.day <= 7 and not is_working_sat)

            # Default / initial values
            punch_in = "-"
            punch_out = "-"
            break_str = "-"
            work_duration = "00:00"
            ot_str = "+00:00"
            ut_str = "-00:00"
            status_badge = "Absent"
            late_flag = "-"
            remarks = ""
            row_fill = FILL_WHITE if (emp_row % 2 == 0) else FILL_ALT
            is_late = False
            ot_mins = 0
            ut_mins = 0

            cin = rec.get("check_in") or rec.get("punch_in") if rec else None
            cout = rec.get("check_out") or rec.get("punch_out") if rec else None

            if rec and cin:
                work_duration = rec.get("work_duration_formatted") or "00:00"
                ot_str = rec.get("overtime_formatted") or "+00:00"
                ut_str = rec.get("undertime_formatted") or "-00:00"
                ot_mins = int(rec.get("overtime_minutes") or round(float(rec.get("overtime_hours", 0.0)) * 60))
                ut_mins = int(rec.get("undertime_minutes") or round(float(rec.get("undertime_hours", 0.0)) * 60))

                if (not rec.get("work_duration_formatted") or rec.get("work_duration_formatted") == "00:00") and cout:
                    attendance_service.apply_daily_calc_fields(rec, rec_shift)
                    work_duration = rec.get("work_duration_formatted", "00:00")
                    ot_str = rec.get("overtime_formatted", "+00:00")
                    ut_str = rec.get("undertime_formatted", "-00:00")
                    ot_mins = int(rec.get("overtime_minutes") or round(float(rec.get("overtime_hours", 0.0)) * 60))
                    ut_mins = int(rec.get("undertime_minutes") or round(float(rec.get("undertime_hours", 0.0)) * 60))

                punch_in = cin
                punch_out = cout or "-"
                break_str = f"{int(rec_shift.break_duration_minutes or 0)}m"

                st_val = str(rec.get("status") or "")
                is_wfh_val = bool(rec.get("is_wfh") or st_val == AttendanceStatus.WFH.value or st_val == "wfh")
                is_missed_val = bool(rec.get("is_missed_punch") or st_val == AttendanceStatus.MISSED_PUNCH.value or (cin and not cout and cur_date < today_pkt))
                is_short_val = bool(rec.get("is_short_leave") or st_val == AttendanceStatus.SHORT_LEAVE.value or st_val == "short_leave")
                is_late = bool(rec.get("is_late") or (rec.get("late_minutes", 0) > 0 and rec.get("late_strike", 0) > 0))

                if is_wfh_val:
                    status_badge = "W.F.H"
                    row_fill = FILL_BLUE_WFH
                elif is_missed_val:
                    status_badge = "Missed Punch"
                    row_fill = FILL_LATE_BADGE
                elif is_short_val:
                    status_badge = "Short Leave"
                elif is_late:
                    status_badge = "Late"
                else:
                    status_badge = "Present"

                if is_late:
                    late_mins = rec.get("late_minutes", 0)
                    late_flag = f"Late ({late_mins}m)"

                remarks = rec.get("notes") or ""

            elif leave:
                l_type = leave.get("leave_type")
                if l_type in (LeaveType.WFH.value, "wfh"):
                    status_badge = "W.F.H"
                    row_fill = FILL_BLUE_WFH
                else:
                    status_badge = "On Leave"
                    row_fill = FILL_BLUE_WFH
                remarks = f"{leave.get('leave_type', 'Leave')}: {leave.get('reason') or ''}".strip(": ")

            elif is_holiday:
                status_badge = "Holiday"
                row_fill = FILL_WEEKEND
                remarks = calendar_titles.get(cur_date_str, "Public Holiday")

            elif is_sun:
                status_badge = "Sunday Off"
                row_fill = FILL_WEEKEND

            elif is_first_sat:
                status_badge = "1st Sat Off"
                row_fill = FILL_WEEKEND

            elif cur_date > today_pkt and cur_date.year == today_pkt.year and cur_date.month == today_pkt.month:
                status_badge = "-"
                ut_str = "-00:00"
                ut_mins = 0

            else:
                # Scheduled workday without punch or approved leave
                status_badge = "Absent"
                expected_mins = int(round(float(rec_shift.expected_hours) * 60))
                ut_str = format_minutes_to_hhmm(-expected_mins, show_sign=True)
                ut_mins = expected_mins

            ws_emp.row_dimensions[emp_row].height = 20

            shift_display_name = (
                "Weekly Rest" if is_sun
                else "Monthly Rest" if is_first_sat
                else (rec.get("shift_name") or rec_shift.name if (rec and cin) else rec_shift.name)
            )

            row_cells_spec = [
                (cur_date_str, ALIGN_CENTER, FONT_DATA_REGULAR),
                (day_name, ALIGN_LEFT, FONT_DATA_REGULAR),
                (shift_display_name, ALIGN_LEFT, FONT_DATA_REGULAR),
                (punch_in, ALIGN_CENTER, FONT_DATA_REGULAR),
                (punch_out, ALIGN_CENTER, FONT_DATA_REGULAR),
                (break_str, ALIGN_CENTER, FONT_DATA_MUTED),
                (work_duration, ALIGN_RIGHT, FONT_DATA_BOLD if work_duration != "00:00" else FONT_DATA_MUTED),
                (ot_str, ALIGN_RIGHT, FONT_OVERTIME if ot_mins > 0 else FONT_DATA_MUTED),
                (ut_str, ALIGN_RIGHT, FONT_UNDERTIME if ut_mins > 0 else FONT_DATA_MUTED),
                (status_badge, ALIGN_CENTER, FONT_LATE if status_badge in ("Late", "Missed Punch", "Absent") else FONT_DATA_BOLD),
                (late_flag, ALIGN_CENTER, FONT_LATE if late_flag != "-" else FONT_DATA_MUTED),
                (remarks, ALIGN_LEFT, FONT_DATA_REGULAR),
            ]

            for col_idx, (val, align, font) in enumerate(row_cells_spec, start=1):
                c = ws_emp.cell(row=emp_row, column=col_idx, value=val)
                c.font = font
                c.fill = row_fill
                c.border = BORDER_CELL
                c.alignment = align

            # Cell override for late flag badge
            if is_late:
                ws_emp.cell(row=emp_row, column=11).fill = FILL_LATE_BADGE

            emp_row += 1

        # Footer Row for Individual Employee Timesheet
        ws_emp.row_dimensions[emp_row].height = 24
        ws_emp.merge_cells(f"A{emp_row}:E{emp_row}")
        ws_emp[f"A{emp_row}"] = "TOTALS / MONTHLY SUMMARY"
        ws_emp[f"A{emp_row}"].font = FONT_DATA_BOLD
        ws_emp[f"A{emp_row}"].alignment = ALIGN_LEFT

        total_work_fmt = agg.total_work_hours_formatted if agg else "00:00"
        total_ot_fmt = agg.overtime_formatted if agg else "+00:00"
        total_ut_fmt = agg.undertime_formatted if agg else "-00:00"
        total_net_fmt = agg.net_variance_formatted if agg else "+00:00"
        punct_score_val = f"{agg.punctuality_percentage:.1f}%" if agg else "100.0%"
        bonus_val = agg.bonus_recommendation.value if (agg and hasattr(agg.bonus_recommendation, "value")) else "Eligible"

        ws_emp.cell(row=emp_row, column=6, value="-").font = FONT_DATA_BOLD
        ws_emp.cell(row=emp_row, column=7, value=total_work_fmt).font = FONT_DATA_BOLD
        ws_emp.cell(row=emp_row, column=8, value=total_ot_fmt).font = FONT_OVERTIME
        ws_emp.cell(row=emp_row, column=9, value=total_ut_fmt).font = FONT_UNDERTIME
        ws_emp.cell(row=emp_row, column=10, value=f"Score: {punct_score_val}").font = FONT_DATA_BOLD
        ws_emp.cell(row=emp_row, column=11, value=f"Net: {total_net_fmt}").font = FONT_DATA_BOLD
        ws_emp.cell(row=emp_row, column=12, value=f"Bonus: {bonus_val}").font = FONT_DATA_BOLD

        for col_idx in range(1, 13):
            c = ws_emp.cell(row=emp_row, column=col_idx)
            c.fill = FILL_FOOTER
            c.border = BORDER_FOOTER
            if col_idx in [7, 8, 9]:
                c.alignment = ALIGN_RIGHT
            elif col_idx in [10, 11, 12]:
                c.alignment = ALIGN_LEFT
            else:
                c.alignment = ALIGN_CENTER

        auto_fit_columns(ws_emp, min_width=11, padding=3)

    # Save to BytesIO
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    logger.info(f"Generated multi-tab attendance workbook for {year}-{month:02d} ({len(wb.worksheets)} worksheets, {len(buffer.getvalue())} bytes).")
    return buffer
